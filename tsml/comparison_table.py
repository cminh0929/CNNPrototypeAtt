"""
Comparison table generation for CNNProto vs SOTA methods.

This module collects CNNProto results and compares them with SOTA methods
from literature, generating formatted comparison tables.
"""

import json
import pandas as pd
from pathlib import Path
from typing import Dict, List


def collect_cnnproto_results() -> pd.DataFrame:
    """Collect CNNProto results from all evaluated datasets.
    
    Returns:
        DataFrame containing dataset names and accuracies.
    """
    results = []
    results_dir = Path("results")
    
    if not results_dir.exists():
        print(f"Warning: {results_dir} does not exist!")
        return pd.DataFrame()
    
    for dataset_dir in results_dir.iterdir():
        if not dataset_dir.is_dir():
            continue
        
        result_file = dataset_dir / "evaluation" / "results.json"
        if result_file.exists():
            try:
                with open(result_file, 'r') as f:
                    data = json.load(f)
                    
                    # [ĐÃ SỬA] Đã bỏ điều kiện lọc accuracy = 1
                    # Tất cả kết quả sẽ được thêm vào danh sách
                    results.append({
                        'Dataset': data['dataset'],
                        'CNNProto': round(data['accuracy'], 4),
                        'Train_Size': data.get('num_samples', 0),
                    })
            except Exception as e:
                print(f"Error reading {result_file}: {e}")
    
    if not results:
        print("No results found! Run 'python main_eval.py --all' first.")
        return pd.DataFrame()
    
    return pd.DataFrame(results).sort_values('Dataset')


def load_sota_results() -> Dict[str, Dict[str, float]]:
    """Load SOTA results from JSON file.
    
    Returns:
        Dictionary mapping dataset names to method accuracies.
    """
    sota_file = Path(__file__).parent / "sota_results.json"
    
    if not sota_file.exists():
        print(f"Warning: {sota_file} not found! Using empty SOTA results.")
        return {}
    
    try:
        with open(sota_file, 'r') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading SOTA results: {e}")
        return {}


def add_sota_results(df: pd.DataFrame) -> pd.DataFrame:
    """Add SOTA results from JSON file.
    """
    # Load SOTA results from JSON
    sota_results = load_sota_results()
    
    # Add SOTA results to dataframe
    for dataset, methods in sota_results.items():
        for method, acc in methods.items():
            idx = df[df['Dataset'] == dataset].index
            if len(idx) > 0:
                df.loc[idx, method] = acc
    
    return df


def calculate_ranks(df: pd.DataFrame, methods: List[str]) -> pd.DataFrame:
    """Calculate rank for each method on each dataset."""
    # Calculate rank (1 = best, higher = worse)
    # Only calculate rank for columns that actually exist in the dataframe
    existing_methods = [m for m in methods if m in df.columns]
    
    if not existing_methods:
        return df

    ranks = df[existing_methods].rank(axis=1, ascending=False, method='average')
    
    for method in existing_methods:
        df[f'{method}_Rank'] = ranks[method]
    
    return df


def create_comparison_table() -> pd.DataFrame:
    """Create complete comparison table."""
    print("\n" + "="*80)
    print("Creating Comparison Table...")
    print("="*80)
    
    # 1. Collect CNNProto results
    print("\n[1/4] Collecting CNNProto results...")
    df = collect_cnnproto_results()
    
    if df.empty:
        return df
    
    print(f"Found results for {len(df)} datasets")
    
    # 2. Add SOTA results
    print("\n[2/4] Adding SOTA results from JSON...")
    df = add_sota_results(df)
    
    # 3. Calculate ranks
    print("\n[3/4] Calculating ranks...")
    
    # Danh sách method theo file JSON mới của bạn
    sota_methods = ['HC2', 'MultiROCKET', 'InceptionTime', 'ResNet', 'TS-CHIEF', '1-NN-DTW']
    # Luôn giữ CNNProto ở đầu
    methods = ['CNNProto'] + sota_methods
    
    available_methods = [m for m in methods if m in df.columns]
    
    if len(available_methods) > 1:
        df = calculate_ranks(df, available_methods)
    
    # 4. Save results
    print("\n[4/5] Saving results...")
    
    # Save CSV
    output_csv = Path("results") / "comparison_table.csv"
    df.to_csv(output_csv, index=False)
    print(f"Saved CSV to: {output_csv}")
    
    # 5. Save Excel with formatting
    print("\n[5/5] Creating Excel file with formatting...")
    output_excel = Path("results") / "comparison_table.xlsx"
    save_to_excel(df, output_excel, available_methods)
    print(f"Saved Excel to: {output_excel}")
    
    return df


def save_to_excel(df: pd.DataFrame, output_file: Path, methods: List[str]) -> None:
    """Save DataFrame to Excel with formatting."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("\nWarning: openpyxl not installed. Installing...")
        import subprocess
        subprocess.check_call(['pip', 'install', 'openpyxl'])
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    
    # Select columns to export
    display_cols = ['Dataset'] + methods
    rank_cols = [f'{m}_Rank' for m in methods if f'{m}_Rank' in df.columns]
    
    # Filter columns that actually exist
    existing_cols = [c for c in display_cols + rank_cols if c in df.columns]
    export_df = df[existing_cols].copy()
    
    # Export to Excel
    export_df.to_excel(output_file, index=False, sheet_name='Comparison')
    
    # Load workbook for formatting
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    best_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    best_font = Font(color="006100", bold=True)
    
    worst_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    worst_font = Font(color="9C0006")
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Format header row
    for col_idx, col in enumerate(export_df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Format data rows
    for row_idx in range(2, len(export_df) + 2):
        for col_idx in range(1, len(export_df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left', vertical='center')
            
            # Format accuracy columns (highlight best/worst per row)
            col_name = export_df.columns[col_idx - 1]
            if col_name in methods:
                # Get values for this row across all methods
                row_values = [export_df.iloc[row_idx - 2][m] for m in methods 
                              if m in export_df.columns and pd.notna(export_df.iloc[row_idx - 2][m])]
                if row_values:
                    max_val = max(row_values)
                    min_val = min(row_values)
                    cell_val = export_df.iloc[row_idx - 2][col_name]
                    
                    if pd.notna(cell_val):
                        # Highlight best (green)
                        if cell_val == max_val:
                            cell.fill = best_fill
                            cell.font = best_font
                        # Highlight worst (red)
                        elif cell_val == min_val and len(row_values) > 1:
                            cell.fill = worst_fill
                            cell.font = worst_font
    
    # Auto-adjust column widths
    for col_idx, col in enumerate(export_df.columns, 1):
        column_letter = get_column_letter(col_idx)
        max_length = len(str(col))
        for row_idx in range(2, len(export_df) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Save
    wb.save(output_file)


def print_summary(df: pd.DataFrame) -> None:
    """Print summary statistics."""
    if df.empty:
        return
    
    print("\n" + "="*80)
    print("COMPARISON TABLE: CNNProto vs SOTA Methods")
    print("="*80)
    
    sota_methods = ['HC2', 'MultiROCKET', 'InceptionTime', 'ResNet', 'TS-CHIEF', '1-NN-DTW']
    
    # Select columns to display
    display_cols = ['Dataset', 'CNNProto']
    for col in sota_methods:
        if col in df.columns:
            display_cols.append(col)
    
    # Print table
    print(df[display_cols].to_string(index=False))
    
    # Print statistics
    print("\n" + "="*80)
    print("STATISTICS")
    print("="*80)
    
    # Average accuracy
    print("\nAverage Accuracy:")
    for col in display_cols[1:]:  # Skip 'Dataset'
        if col in df.columns:
            avg = df[col].mean()
            print(f"  {col:20s}: {avg:.4f}")
    
    # Average rank
    if 'CNNProto_Rank' in df.columns:
        print(f"\nAverage Rank:")
        for col in display_cols[1:]:
            rank_col = f"{col}_Rank"
            if rank_col in df.columns:
                avg_rank = df[rank_col].mean()
                print(f"  {col:20s}: {avg_rank:.2f}")
    
    # Win/Tie/Loss
    print("\nWin/Tie/Loss vs CNNProto:")
    for method in sota_methods:
        if method in df.columns:
            wins = (df['CNNProto'] > df[method]).sum()
            ties = (df['CNNProto'] == df[method]).sum()
            losses = (df['CNNProto'] < df[method]).sum()
            print(f"  {method:20s}: {wins:2d} / {ties:2d} / {losses:2d}")
    
    print("="*80)


def main() -> None:
    """Main entry point for comparison table generation."""
    df = create_comparison_table()
    
    if not df.empty:
        print_summary(df)
        
        print("\n[OK] Done! Results saved:")
        print("  • CSV:   results/comparison_table.csv")
        print("  • Excel: results/comparison_table.xlsx (with formatting)")
    else:
        print("\n[x] No results found!")
        print("\nPlease run evaluations first:")
        print("  python main_eval.py --all")

if __name__ == "__main__":
    main()